#!/usr/bin/env python3
"""Create an optional Excel tracker for the homeschool learning agent."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

BANDS = ["Emerging", "Developing", "Secure", "Advanced"]


def style_header(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="D9EAF7")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="999999")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def add_band_validation(ws, start_row: int, end_row: int, col: str):
    dv = DataValidation(type="list", formula1='"' + ",".join(BANDS) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}{start_row}:{col}{end_row}")


def build_sheet(ws, learner_name: str, age: int, country: str, weeks: int):
    ws.title = "Tracker"
    headers = ["Field", "Value"]
    for i, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=i, value=h))
    values = [
        ("Learner name", learner_name),
        ("Age", age),
        ("Country", country),
        ("Weeks", weeks),
        ("Priorities", ""),
        ("Concerns", ""),
    ]
    for r, (k, v) in enumerate(values, start=2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    start = 10
    cols = ["Domain", "Subskill", "Band", "Confidence", "Notes", "Next step"]
    for i, h in enumerate(cols, start=1):
        style_header(ws.cell(row=start, column=i, value=h))
    domains = ["Core language", "Maths", "Science"]
    row = start + 1
    for domain in domains:
        for _ in range(4):
            ws.cell(row=row, column=1, value=domain)
            row += 1
    add_band_validation(ws, start + 1, row - 1, "C")
    set_widths(ws, {"A": 18, "B": 24, "C": 14, "D": 12, "E": 30, "F": 24})


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--age", type=int, required=True)
    parser.add_argument("--country", required=True)
    parser.add_argument("--learner-name", default="Learner")
    parser.add_argument("--weeks", type=int, default=12)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    build_sheet(wb.active, args.learner_name, args.age, args.country, args.weeks)
    wb.save(output)
    print(f"Created: {output}")


if __name__ == "__main__":
    main()
